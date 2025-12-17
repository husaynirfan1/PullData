"""Excel formatter for PullData output.

Supports both openpyxl (feature-rich) and xlsxwriter (fast) backends.
Provides automatic styling, multi-sheet support, and table formatting.
"""

from io import BytesIO
from typing import Any, Dict, List, Literal, Optional

from pulldata.synthesis.base import FormatConfig, FormatterError, OutputData, OutputFormatter

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter

    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False


class ExcelFormatter(OutputFormatter):
    """Excel formatter using openpyxl or xlsxwriter.
    
    Features:
    - Auto-styling (headers, filters, freeze panes)
    - Multi-sheet support
    - Table structure preservation
    - Cell formatting (bold, colors, borders)
    
    Example:
        >>> formatter = ExcelFormatter()
        >>> data = OutputData(title="Report", content="Summary", tables=[...])
        >>> formatter.save(data, "output.xlsx")
    """

    def __init__(
        self,
        config: Optional[FormatConfig] = None,
        backend: Literal["openpyxl", "xlsxwriter", "auto"] = "auto",
        enable_auto_styling: bool = True,
        enable_filters: bool = True,
        freeze_header: bool = True,
    ):
        """Initialize Excel formatter.
        
        Args:
            config: Formatter configuration
            backend: Excel library to use ('openpyxl', 'xlsxwriter', or 'auto')
            enable_auto_styling: Apply automatic styling to tables
            enable_filters: Add autofilter to tables
            freeze_header: Freeze top row in sheets with tables
        """
        super().__init__(config)

        # Determine backend
        if backend == "auto":
            if OPENPYXL_AVAILABLE:
                self.backend = "openpyxl"
            elif XLSXWRITER_AVAILABLE:
                self.backend = "xlsxwriter"
            else:
                raise FormatterError(
                    "No Excel backend available. Install openpyxl or xlsxwriter",
                    formatter="Excel",
                    details={"install": "pip install openpyxl xlsxwriter"},
                )
        else:
            self.backend = backend
            if backend == "openpyxl" and not OPENPYXL_AVAILABLE:
                raise FormatterError(
                    "openpyxl not available",
                    formatter="Excel",
                    details={"install": "pip install openpyxl"},
                )
            if backend == "xlsxwriter" and not XLSXWRITER_AVAILABLE:
                raise FormatterError(
                    "xlsxwriter not available",
                    formatter="Excel",
                    details={"install": "pip install xlsxwriter"},
                )

        self.enable_auto_styling = enable_auto_styling
        self.enable_filters = enable_filters
        self.freeze_header = freeze_header

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    @property
    def format_name(self) -> str:
        return "Excel"

    def format(self, data: OutputData) -> bytes:
        """Format data as Excel file.
        
        Args:
            data: Data to format
            
        Returns:
            Excel file as bytes
        """
        if self.backend == "openpyxl":
            return self._format_openpyxl(data)
        else:
            return self._format_xlsxwriter(data)

    def _format_openpyxl(self, data: OutputData) -> bytes:
        """Format using openpyxl backend."""
        if not OPENPYXL_AVAILABLE:
            raise FormatterError(
                "openpyxl not available",
                formatter="Excel",
                details={"install": "pip install openpyxl"},
            )

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Summary sheet
        self._create_summary_sheet_openpyxl(wb, data)

        # Table sheets
        if data.tables:
            for i, table_data in enumerate(data.tables):
                sheet_name = table_data.get("name", f"Table_{i+1}")
                self._create_table_sheet_openpyxl(wb, sheet_name, table_data)

        # Sources sheet (if requested)
        if self.config.include_sources and data.sources:
            self._create_sources_sheet_openpyxl(wb, data.sources)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _create_summary_sheet_openpyxl(self, wb: Any, data: OutputData) -> None:
        """Create summary sheet with title and content."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = data.title
        ws["A1"].font = Font(size=16, bold=True)
        ws.row_dimensions[1].height = 24

        # Content
        ws["A3"] = data.content
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

        # Metadata
        if self.config.include_metadata and data.metadata:
            row = 5
            ws[f"A{row}"] = "Metadata"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            for key, value in data.metadata.items():
                ws[f"A{row}"] = str(key)
                ws[f"B{row}"] = str(value)
                row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

    def _create_table_sheet_openpyxl(
        self, wb: Any, sheet_name: str, table_data: Dict[str, Any]
    ) -> None:
        """Create a sheet with table data."""
        ws = wb.create_sheet(sheet_name)

        # Extract table structure
        headers = table_data.get("headers", [])
        rows = table_data.get("rows", [])

        if not headers and not rows:
            ws["A1"] = "No table data"
            return

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            if self.enable_auto_styling:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add table formatting
        if headers and rows and self.enable_auto_styling:
            table_range = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
            table = Table(displayName=sheet_name.replace(" ", "_"), ref=table_range)

            # Add a default style
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        # Add autofilter
        if self.enable_filters and headers:
            ws.auto_filter.ref = ws.dimensions

        # Freeze header row
        if self.freeze_header and headers:
            ws.freeze_panes = "A2"

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_sources_sheet_openpyxl(self, wb: Any, sources: List[Dict[str, Any]]) -> None:
        """Create sources/citations sheet."""
        ws = wb.create_sheet("Sources")

        # Headers
        headers = ["Document", "Page", "Chunk", "Relevance"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if self.enable_auto_styling:
                cell.font = Font(bold=True)

        # Data
        for row_idx, source in enumerate(sources, 2):
            ws.cell(row=row_idx, column=1, value=source.get("document_id", ""))
            ws.cell(row=row_idx, column=2, value=source.get("page_number", ""))
            ws.cell(row=row_idx, column=3, value=source.get("chunk_id", ""))
            ws.cell(row=row_idx, column=4, value=source.get("score", ""))

        # Auto-adjust column widths
        for col_idx in range(1, 5):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def _format_xlsxwriter(self, data: OutputData) -> bytes:
        """Format using xlsxwriter backend."""
        if not XLSXWRITER_AVAILABLE:
            raise FormatterError(
                "xlsxwriter not available",
                formatter="Excel",
                details={"install": "pip install xlsxwriter"},
            )

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})

        # Define formats
        title_format = workbook.add_format({"bold": True, "font_size": 16})
        header_format = workbook.add_format(
            {"bold": True, "bg_color": "#4472C4", "font_color": "white", "align": "center"}
        )
        wrap_format = workbook.add_format({"text_wrap": True, "valign": "top"})

        # Summary sheet
        summary_ws = workbook.add_worksheet("Summary")
        summary_ws.write("A1", data.title, title_format)
        summary_ws.set_row(0, 24)
        summary_ws.write("A3", data.content, wrap_format)
        summary_ws.set_column("A:A", 30)
        summary_ws.set_column("B:B", 50)

        # Metadata
        if self.config.include_metadata and data.metadata:
            row = 4
            summary_ws.write(row, 0, "Metadata", header_format)
            row += 1
            for key, value in data.metadata.items():
                summary_ws.write(row, 0, str(key))
                summary_ws.write(row, 1, str(value))
                row += 1

        # Table sheets
        if data.tables:
            for i, table_data in enumerate(data.tables):
                sheet_name = table_data.get("name", f"Table_{i+1}")
                self._create_table_sheet_xlsxwriter(
                    workbook, sheet_name, table_data, header_format
                )

        # Sources sheet
        if self.config.include_sources and data.sources:
            self._create_sources_sheet_xlsxwriter(workbook, data.sources, header_format)

        workbook.close()
        output.seek(0)
        return output.read()

    def _create_table_sheet_xlsxwriter(
        self, workbook: Any, sheet_name: str, table_data: Dict[str, Any], header_format: Any
    ) -> None:
        """Create table sheet with xlsxwriter."""
        worksheet = workbook.add_worksheet(sheet_name)

        headers = table_data.get("headers", [])
        rows = table_data.get("rows", [])

        if not headers and not rows:
            worksheet.write("A1", "No table data")
            return

        # Write headers
        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, str(header), header_format if self.enable_auto_styling else None)

        # Write data
        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row):
                worksheet.write(row_idx, col_idx, value)

        # Add autofilter
        if self.enable_filters and headers:
            worksheet.autofilter(0, 0, len(rows), len(headers) - 1)

        # Freeze header
        if self.freeze_header and headers:
            worksheet.freeze_panes(1, 0)

        # Set column widths
        for col_idx in range(len(headers)):
            worksheet.set_column(col_idx, col_idx, 15)

    def _create_sources_sheet_xlsxwriter(
        self, workbook: Any, sources: List[Dict[str, Any]], header_format: Any
    ) -> None:
        """Create sources sheet with xlsxwriter."""
        worksheet = workbook.add_worksheet("Sources")

        headers = ["Document", "Page", "Chunk", "Relevance"]
        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, header, header_format if self.enable_auto_styling else None)

        for row_idx, source in enumerate(sources, 1):
            worksheet.write(row_idx, 0, source.get("document_id", ""))
            worksheet.write(row_idx, 1, source.get("page_number", ""))
            worksheet.write(row_idx, 2, source.get("chunk_id", ""))
            worksheet.write(row_idx, 3, source.get("score", ""))

        for col_idx in range(4):
            worksheet.set_column(col_idx, col_idx, 20)
